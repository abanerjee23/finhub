from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from pydantic import BaseModel

from cfin_agents.eval_cases import SummaryCase, load_summary_cases, workflow_summary
from cfin_agents.summary_judge import grade_agent_summary
from cfin_agents.workflow import run_document_workflow

MODEL_OUTPUTS_PATH = (
    Path(__file__).resolve().parents[2] / "evals" / "model_outputs.jsonl"
)


class SummaryEvalRecord(BaseModel):
    run_id: str
    timestamp: str
    case_id: str
    document_id: str
    execution_mode: str
    expected_status: str
    actual_status: str
    overall_pass: bool
    accuracy_score: int | None = None
    actionability_score: int | None = None
    audience_fit_score: int | None = None
    conciseness_score: int | None = None
    judge_reasoning: str | None = None
    agent_summary: str | None = None
    error: str | None = None


def append_summary_eval_record(record: SummaryEvalRecord, path: Path = MODEL_OUTPUTS_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(record.model_dump_json())
        file.write("\n")


def load_summary_eval_records(path: Path = MODEL_OUTPUTS_PATH) -> list[SummaryEvalRecord]:
    if not path.exists():
        return []

    records: list[SummaryEvalRecord] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            records.append(SummaryEvalRecord.model_validate(json.loads(line)))
    return records


def latest_run_records(path: Path = MODEL_OUTPUTS_PATH) -> list[SummaryEvalRecord]:
    records = load_summary_eval_records(path)
    if not records:
        return []
    latest_run_id = records[-1].run_id
    return [record for record in records if record.run_id == latest_run_id]


def run_summary_eval_batch(
    *,
    run_id: str | None = None,
    skip_judge: bool = False,
) -> tuple[str, list[SummaryEvalRecord]]:
    batch_id = run_id or datetime.now(UTC).strftime("eval-%Y%m%dT%H%M%SZ")
    timestamp = datetime.now(UTC).isoformat()
    records: list[SummaryEvalRecord] = []

    for case in load_summary_cases():
        record = _evaluate_summary_case(
            case=case,
            run_id=batch_id,
            timestamp=timestamp,
            skip_judge=skip_judge,
        )
        append_summary_eval_record(record)
        records.append(record)

    return batch_id, records


def _evaluate_summary_case(
    *,
    case: SummaryCase,
    run_id: str,
    timestamp: str,
    skip_judge: bool,
) -> SummaryEvalRecord:
    base = SummaryEvalRecord(
        run_id=run_id,
        timestamp=timestamp,
        case_id=case.id,
        document_id=case.document_id,
        execution_mode="unknown",
        expected_status=case.golden.expected_status,
        actual_status="unknown",
        overall_pass=False,
    )

    try:
        run = run_document_workflow(
            document_id=case.document_id,
            approve=case.approve,
            force_deterministic=False,
        )
        structured = workflow_summary(run)
        agent_summary = run.agent_summary or ""
        base.execution_mode = run.execution_mode
        base.actual_status = str(structured.get("status", "unknown"))
        base.agent_summary = agent_summary or None

        smoke_pass = (
            base.actual_status == case.golden.expected_status
            and structured.get("reason_code") == case.golden.expected_reason_code
            and structured.get("action") == case.golden.expected_action
            and bool(agent_summary)
        )

        if skip_judge:
            base.overall_pass = smoke_pass
            if not smoke_pass:
                base.error = "Smoke checks failed or agent_summary missing."
            return base

        if not agent_summary:
            base.error = "agent_summary missing"
            return base

        grade = grade_agent_summary(case, agent_summary, structured)
        base.accuracy_score = int(grade["accuracy_score"])
        base.actionability_score = int(grade["actionability_score"])
        base.audience_fit_score = int(grade.get("audience_fit_score", 0))
        base.conciseness_score = int(grade.get("conciseness_score", 0))
        base.judge_reasoning = str(grade.get("reasoning", ""))
        base.overall_pass = bool(grade["overall_pass"]) and smoke_pass
        if not base.overall_pass and smoke_pass is False:
            base.error = "Structured smoke checks failed."
        return base
    except Exception as exc:  # noqa: BLE001
        base.error = str(exc)
        return base


def summarize_records(records: list[SummaryEvalRecord]) -> dict[str, Any]:
    total = len(records)
    passed = sum(1 for record in records if record.overall_pass)
    return {
        "total": total,
        "passed": passed,
        "failed": total - passed,
        "pass_rate": round(passed / total, 3) if total else 0.0,
    }


def export_records_to_excel(
    records: list[SummaryEvalRecord],
    workbook_path: Path,
    sheet_name: str = "Model_Outputs",
) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install dev dependencies."
        ) from exc

    headers = [
        "run_id",
        "timestamp",
        "case_id",
        "document_id",
        "execution_mode",
        "expected_status",
        "actual_status",
        "overall_pass",
        "accuracy_score",
        "actionability_score",
        "audience_fit_score",
        "conciseness_score",
        "judge_reasoning",
        "agent_summary",
        "error",
    ]

    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
        sheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook.create_sheet(sheet_name)
        )
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    if sheet.max_row == 0 or sheet.cell(1, 1).value != "run_id":
        sheet.append(headers)

    for record in records:
        sheet.append([getattr(record, header) for header in headers])

    workbook.save(workbook_path)
